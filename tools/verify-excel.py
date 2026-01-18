#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""验证Excel文件"""

import sys
import io
import openpyxl

# 设置Windows控制台输出为UTF-8
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('完美账单.xlsx')
ws = wb['交易明细']

print("=" * 100)
print("Excel数据验证")
print("=" * 100)
print(f"\n{'序号':<5} {'交易日期':<12} {'记账日期':<12} {'交易说明':<25} {'金额':<10} {'收支':<6} {'类型':<8} {'卡号':<6}")
print("-" * 100)

for i in range(2, min(22, ws.max_row + 1)):
    seq = ws.cell(i, 1).value
    trans_date = ws.cell(i, 2).value or ""
    post_date = ws.cell(i, 3).value or ""
    desc = str(ws.cell(i, 4).value or "")[:23]
    amount = ws.cell(i, 5).value
    income_expense = ws.cell(i, 6).value or ""
    trans_type = ws.cell(i, 7).value or ""
    card = ws.cell(i, 8).value or ""
    
    if seq:
        print(f"{seq:<5} {trans_date:<12} {post_date:<12} {desc:<25} {amount:>10.2f} {income_expense:<6} {trans_type:<8} {card:<6}")

# 显示汇总
print("\n" + "=" * 100)
print("汇总信息:")
summary_row = ws.max_row - 3
print(f"合计(收入-支出): {ws.cell(summary_row, 5).value:.2f}")
print(f"总支出: {ws.cell(summary_row + 1, 5).value:.2f}")
print(f"总收入: {ws.cell(summary_row + 2, 5).value:.2f}")
print(f"还款金额(不计入收支): {ws.cell(summary_row + 3, 5).value:.2f}")
print("=" * 100)
