#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PDF账单转Excel工具
支持多种PDF账单格式,自动识别并提取交易记录
"""

import sys
import re
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any
import argparse
import io

# 设置Windows控制台输出为UTF-8
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

try:
    import pdfplumber
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("[错误] 缺少必要的库,请先安装:")
    print("pip install pdfplumber openpyxl")
    sys.exit(1)


class PDFBillConverter:
    """PDF账单转换器"""
    
    def __init__(self, pdf_path: str):
        self.pdf_path = Path(pdf_path)
        if not self.pdf_path.exists():
            raise FileNotFoundError(f"PDF文件不存在: {pdf_path}")
        
        self.transactions: List[Dict[str, Any]] = []
        self.bill_info: Dict[str, str] = {}
    
    def extract_text(self) -> str:
        """提取PDF文本内容"""
        print(f"[读取] 正在读取PDF: {self.pdf_path.name}")
        
        full_text = ""
        with pdfplumber.open(self.pdf_path) as pdf:
            print(f"   总页数: {len(pdf.pages)}")
            for i, page in enumerate(pdf.pages, 1):
                text = page.extract_text()
                if text:
                    full_text += text + "\n"
                print(f"   已处理: {i}/{len(pdf.pages)} 页", end="\r")
        
        print(f"\n[完成] PDF读取完成,共提取 {len(full_text)} 个字符")
        return full_text
    
    def extract_tables(self) -> List[List[List[str]]]:
        """提取PDF中的表格数据"""
        print(f"[表格] 正在提取表格数据...")
        
        all_tables = []
        with pdfplumber.open(self.pdf_path) as pdf:
            for i, page in enumerate(pdf.pages, 1):
                tables = page.extract_tables()
                if tables:
                    all_tables.extend(tables)
                    print(f"   第{i}页找到 {len(tables)} 个表格")
        
        print(f"[完成] 共提取 {len(all_tables)} 个表格")
        return all_tables
    
    def parse_credit_card_bill(self, text: str, tables: List[List[List[str]]]):
        """解析信用卡账单"""
        print("[解析] 正在解析信用卡账单...")
        
        # 提取账单基本信息
        self._extract_bill_info(text)
        
        # 方法1: 从表格中提取
        header_found = False
        
        for table in tables:
            if not table or len(table) == 0:
                continue
            
            first_row = table[0]
            
            # 检查是否是表头
            if self._is_header_row(first_row):
                header_found = True
                print(f"[表头] 找到表头: {first_row}")
                
                # 如果表格有多行,处理后续行
                if len(table) > 1:
                    for row in table[1:]:
                        transaction = self._parse_transaction_row_cmb(row)
                        if transaction:
                            self.transactions.append(transaction)
            
            # 如果已经找到表头,后续的单行表格都是交易记录
            elif header_found and len(table) == 1:
                transaction = self._parse_transaction_row_cmb(first_row)
                if transaction:
                    self.transactions.append(transaction)
        
        print(f"[表格] 从表格中提取了 {len(self.transactions)} 条记录")
        
        # 方法2: 从文本中提取(补充遗漏的记录)
        text_transactions = self._parse_transactions_from_text(text)
        
        # 去重:检查文本提取的记录是否已存在
        existing_keys = set()
        for t in self.transactions:
            key = f"{t['交易日期']}_{t['交易说明']}_{t['金额']}"
            existing_keys.add(key)
        
        added_count = 0
        for t in text_transactions:
            key = f"{t['交易日期']}_{t['交易说明']}_{t['金额']}"
            if key not in existing_keys:
                self.transactions.append(t)
                existing_keys.add(key)
                added_count += 1
        
        if added_count > 0:
            print(f"[文本] 从文本中补充了 {added_count} 条遗漏的记录")
        
        print(f"[完成] 解析完成,共找到 {len(self.transactions)} 条交易记录")
    
    def _extract_bill_info(self, text: str):
        """提取账单基本信息"""
        patterns = {
            '账单月份': r'账单(?:月份|周期)[：:]\s*(\d{4}[-/年]\d{1,2}(?:[-/月]\d{1,2})?)',
            '卡号': r'卡号[：:]\s*(\d{4}\s*\*+\s*\d{4})',
            '持卡人': r'持卡人[：:]\s*([^\n\r]+)',
            '本期应还': r'本期应还(?:金额)?[：:]\s*[¥￥]?\s*([\d,]+\.?\d*)',
            '最后还款日': r'最后还款日[：:]\s*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2})',
        }
        
        for key, pattern in patterns.items():
            match = re.search(pattern, text)
            if match:
                self.bill_info[key] = match.group(1).strip()
    
    
    def _is_header_row(self, row: List[str]) -> bool:
        """判断是否是表头行"""
        if not row or len(row) < 3:
            return False
        
        row_text = ' '.join([str(cell) for cell in row if cell]).lower()
        # 招商银行信用卡账单表头特征
        keywords = ['交易日', '记账日', '交易摘要', '人民币金额', 'trans', 'post', 'description', 'amount']
        return sum(1 for kw in keywords if kw in row_text) >= 3
    
    def _parse_transaction_row_cmb(self, row: List[str]) -> Dict[str, Any]:
        """解析招商银行信用卡账单的交易行
        
        格式: [交易日, 记账日, 交易摘要, 人民币金额, 卡号末四位, 交易地金额]
        还款: ['', '11/27', '自动还款', '-1,731.52', '3686', '-1,731.52']  # 还款不计入收支
        退款: ['11/23', '11/24', '支付宝-拼多多平台商户', '-25.20', '3686', '-25.20(CN)']  # 退款=收入
        消费: ['11/23', '11/24', '支付宝-拼多多平台商户', '25.20', '3686', '25.20(CN)']  # 消费=支出
        """
        if not row or len(row) < 3:
            return None
        
        try:
            # 检查第一列是否为空(还款类交易)
            if not row[0] or str(row[0]).strip() == '':
                # 还款格式: ['', 交易日, 交易摘要, 金额, 卡号, 交易地金额]
                trans_date = row[1] if len(row) > 1 else ''
                post_date = trans_date  # 还款没有单独的记账日
                description = row[2] if len(row) > 2 else ''
                amount_str = row[3] if len(row) > 3 else '0'
                card_num = row[4] if len(row) > 4 else ''
                original_amount = row[5] if len(row) > 5 else amount_str
            elif len(row) == 6:
                # 标准格式: [交易日, 记账日, 交易摘要, 金额, 卡号, 交易地金额]
                trans_date = row[0]
                post_date = row[1]
                description = row[2]
                amount_str = row[3]
                card_num = row[4]
                original_amount = row[5]
            else:
                # 尝试智能识别
                trans_date = row[0] if len(row) > 0 else ''
                description = row[1] if len(row) > 1 else ''
                amount_str = row[2] if len(row) > 2 else '0'
                card_num = row[3] if len(row) > 3 else ''
                post_date = trans_date
                original_amount = amount_str
            
            # 跳过表头行
            if '交易日' in str(trans_date) or 'Trans' in str(trans_date):
                return None
            
            # 解析金额(保留正负号)
            amount = self._parse_amount(amount_str)
            if amount == 0:
                return None
            
            # 解析日期
            parsed_date = self._parse_date_cmb(trans_date)
            
            # 判断交易类型和收支
            # 新逻辑:
            # - 还款: 类型=还款,收支=还款(不计入收支统计)
            # - 退款: 类型=退款,收支=收入
            # - 消费: 类型=消费,收支=支出
            desc = str(description).lower()
            
            if '还款' in desc or 'payment' in desc:
                trans_type = '还款'
                income_or_expense = '还款'  # 还款单独标记,不计入收支
            elif '退款' in desc or 'refund' in desc or (amount < 0 and '还款' not in desc):
                trans_type = '退款'
                income_or_expense = '收入'  # 退款计入收入
            else:
                trans_type = '消费'
                income_or_expense = '支出'  # 消费计入支出
            
            return {
                '交易日期': parsed_date,
                '记账日期': self._parse_date_cmb(post_date),
                '交易说明': str(description).strip(),
                '金额': abs(amount),
                '类型': trans_type,
                '收支': income_or_expense,
                '卡号末四位': str(card_num).strip(),
                '币种': 'CNY'
            }
            
        except Exception as e:
            print(f"[警告] 解析行数据失败: {row}, 错误: {e}")
            return None
    
    def _parse_date_cmb(self, date_str: str) -> str:
        """解析招商银行日期格式 (MM/DD)"""
        if not date_str:
            return ''
        
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
[工具] PDF信用卡账单转Excel
将PDF账单数据提取并保存为Excel文件,支持自动分类和统计
"""

import sys
import io
import argparse
from pathlib import Path
from bill_parser import PDFBillConverter
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 设置Windows控制台输出为UTF-8
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

class ExcelSaver:
    """专门负责Excel保存的类(从Converter中分离出来)"""
    def __init__(self, transactions, bill_info, pdf_path):
        self.transactions = transactions
        self.bill_info = bill_info
        self.pdf_path = Path(pdf_path)

    def save_to_excel(self, output_path: str = None):
        """保存为Excel文件"""
        if not self.transactions:
            print("[警告] 没有找到交易记录,无法生成Excel")
            return
        
        # 确定输出路径
        if not output_path:
            output_path = self.pdf_path.with_suffix('.xlsx')
        
        output_path = Path(output_path)
        
        print(f"[保存] 正在生成Excel文件: {output_path.name}")
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        
        # 创建账单信息sheet
        if self.bill_info:
            self._create_info_sheet(wb)
        
        # 创建交易明细sheet
        self._create_transactions_sheet(wb)
        
        # 删除默认sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # 保存文件
        wb.save(output_path)
        print(f"[完成] Excel文件已保存: {output_path}")
        print(f"   共 {len(self.transactions)} 条交易记录")
        
        return output_path
    
    def _create_info_sheet(self, wb):
        """创建账单信息sheet"""
        ws = wb.create_sheet('账单信息', 0)
        
        # 设置标题样式
        title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        title_font = Font(color='FFFFFF', bold=True, size=12)
        
        # 写入信息
        row = 1
        for key, value in self.bill_info.items():
            ws.cell(row, 1, key).fill = title_fill
            ws.cell(row, 1).font = title_font
            ws.cell(row, 2, value)
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
    
    def _create_transactions_sheet(self, wb):
        """创建交易明细sheet"""
        ws = wb.create_sheet('交易明细', 0 if not self.bill_info else 1)
        
        # 定义表头
        headers = ['序号', '交易日期', '记账日期', '交易说明', '金额', '收支', '类型', '卡号末四位', '币种']
        
        # 设置表头样式
        header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 写入数据
        for idx, transaction in enumerate(self.transactions, 1):
            ws.cell(idx + 1, 1, idx).border = border
            ws.cell(idx + 1, 2, transaction.get('交易日期', '')).border = border
            ws.cell(idx + 1, 3, transaction.get('记账日期', '')).border = border
            ws.cell(idx + 1, 4, transaction.get('交易说明', '')).border = border
            
            # 金额单元格 - 支出为负数,收入为正数
            amount = transaction.get('金额', 0)
            is_expense = transaction.get('收支', '') == '支出'
            display_amount = -amount if is_expense else amount
            
            amount_cell = ws.cell(idx + 1, 5, display_amount)
            amount_cell.border = border
            amount_cell.number_format = '#,##0.00'
            
            # 根据收支设置颜色
            if is_expense:
                amount_cell.font = Font(color='FF0000')  # 红色表示支出
            else:
                amount_cell.font = Font(color='00B050')  # 绿色表示收入
            
            ws.cell(idx + 1, 6, transaction.get('收支', '')).border = border
            ws.cell(idx + 1, 7, transaction.get('类型', '')).border = border
            ws.cell(idx + 1, 8, transaction.get('卡号末四位', '')).border = border
            ws.cell(idx + 1, 9, transaction.get('币种', 'CNY')).border = border
        
        # 添加汇总行
        summary_row = len(self.transactions) + 2
        ws.cell(summary_row, 4, '合计:').font = Font(bold=True)
        ws.cell(summary_row, 4).alignment = Alignment(horizontal='right')
        
        # 计算总支出和总收入(不包括还款)
        total_expense = sum(-t.get('金额', 0) for t in self.transactions if t.get('收支') == '支出')
        total_income = sum(t.get('金额', 0) for t in self.transactions if t.get('收支') == '收入')
        total_repayment = sum(-t.get('金额', 0) for t in self.transactions if t.get('收支') == '还款')
        total_amount = total_income + total_expense  # 不包括还款
        
        total_cell = ws.cell(summary_row, 5, total_amount)
        total_cell.font = Font(bold=True, color='FF0000' if total_amount < 0 else '00B050')
        total_cell.number_format = '#,##0.00'
        
        # 添加统计信息
        ws.cell(summary_row + 1, 4, '总支出:').font = Font(bold=True)
        ws.cell(summary_row + 1, 4).alignment = Alignment(horizontal='right')
        expense_cell = ws.cell(summary_row + 1, 5, total_expense)
        expense_cell.font = Font(bold=True, color='FF0000')
        expense_cell.number_format = '#,##0.00'
        
        ws.cell(summary_row + 2, 4, '总收入:').font = Font(bold=True)
        ws.cell(summary_row + 2, 4).alignment = Alignment(horizontal='right')
        income_cell = ws.cell(summary_row + 2, 5, total_income)
        income_cell.font = Font(bold=True, color='00B050')
        income_cell.number_format = '#,##0.00'
        
        # 添加还款统计(单独显示,不计入收支)
        ws.cell(summary_row + 3, 4, '还款金额:').font = Font(bold=True)
        ws.cell(summary_row + 3, 4).alignment = Alignment(horizontal='right')
        repayment_cell = ws.cell(summary_row + 3, 5, total_repayment)
        repayment_cell.font = Font(bold=True, color='0070C0')  # 蓝色表示还款
        repayment_cell.number_format = '#,##0.00'
        
        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 8
        
        # 冻结首行
        ws.freeze_panes = 'A2'


def main():
    parser = argparse.ArgumentParser(description='PDF账单转Excel工具')
    parser.add_argument('pdf_file', help='PDF账单文件路径')
    parser.add_argument('-o', '--output', help='输出Excel文件路径(可选)')
    
    args = parser.parse_args()
    
    try:
        print("=" * 60)
        print("[工具] PDF账单转Excel工具")
        print("=" * 60)
        
        # 使用核心库创建转换器
        converter = PDFBillConverter(args.pdf_file)
        
        # 提取文本和表格
        text = converter.extract_text()
        tables = converter.extract_tables()
        
        # 解析账单
        converter.parse_credit_card_bill(text, tables)
        
        # 委托 ExcelSaver 保存
        saver = ExcelSaver(converter.transactions, converter.bill_info, args.pdf_file)
        output_file = saver.save_to_excel(args.output)
        
        print("=" * 60)
        print("[完成] 转换完成!")
        print("=" * 60)
        
    except Exception as e:
        print(f"\n[错误] 转换失败: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
