#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""统计交易类型"""

import sys
import io
import openpyxl

# 设置Windows控制台输出为UTF-8
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('完美账单.xlsx')
ws = wb['交易明细']

types = {}
income_expense = {}

for i in range(2, ws.max_row - 2):
    trans_type = ws.cell(i, 7).value  # 类型列
    income_exp = ws.cell(i, 6).value  # 收支列
    
    if trans_type:
        types[trans_type] = types.get(trans_type, 0) + 1
    if income_exp:
        income_expense[income_exp] = income_expense.get(income_exp, 0) + 1

print("=" * 60)
print("交易类型统计:")
print("=" * 60)
for k, v in types.items():
    print(f"{k}: {v}笔")
print(f"\n总计: {sum(types.values())}笔")

print("\n" + "=" * 60)
print("收支统计:")
print("=" * 60)
for k, v in income_expense.items():
    print(f"{k}: {v}笔")
print("=" * 60)
